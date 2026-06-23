"""Export engine — PDF (reportlab) and Excel (openpyxl) generators.

Shared by admin reports export endpoints.
"""

import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


# ── PDF ────────────────────────────────────────────────────────────


def generate_pdf(title: str, columns: list[str], rows: list[list], company: str = "NEOGAS") -> bytes:
    """Generate a landscape PDF table report.

    Args:
        title: Report title
        columns: Column headers
        rows: List of row lists (all values as strings)
        company: Company name for header
    """
    buf = io.BytesIO()
    page_w, page_h = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=(page_w, page_h),
                           leftMargin=15*mm, rightMargin=15*mm,
                           topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=14, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle2', parent=styles['Normal'], fontSize=8, textColor=colors.grey)

    elements = []
    elements.append(Paragraph(f"<b>{company}</b> — {title}", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 5*mm))

    # Build table data
    table_data = [columns] + rows

    # Calculate column widths proportionally
    avail_w = page_w - 30*mm
    col_w = avail_w / len(columns)
    col_widths = [col_w] * len(columns)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()


# ── Excel ──────────────────────────────────────────────────────────


def generate_excel(title: str, columns: list[str], rows: list[list], company: str = "NEOGAS") -> bytes:
    """Generate an Excel workbook with one sheet.

    Args:
        title: Sheet title
        columns: Column headers
        rows: List of row lists
        company: Company name
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Company + date row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value=f"{company} — {title}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color="666666")

    # Headers at row 3
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows starting at row 4
    for row_idx, row in enumerate(rows, 4):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = Font(size=9)

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(columns) + 1):
        max_len = len(columns[col_idx - 1])
        for row in rows:
            cell_val = str(row[col_idx - 1]) if col_idx <= len(row) else ""
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(max_len + 4, 40)

    # Freeze header
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Shift Close Receipt PDF ──────────────────────────────────────────


def generate_shift_receipt_pdf(
    shift_id: int,
    opened_at: str,
    closed_at: str,
    cashier: str,
    opening_cash: float,
    cash_income: float, cash_income_count: int,
    cash_expense: float, cash_expense_count: int,
    cash_deposits: float, cash_deposits_count: int,
    cash_transfers_out: float, cash_transfers_out_count: int,
    cash_transfers_in: float, cash_transfers_in_count: int,
    cash_safe_drops: float, cash_safe_drops_count: int,
    sales_cash: float, sales_cash_count: int,
    surplus: float, shortage: float,
    non_cash_sales: list[dict],
    total_sales: float,
    dispatch_count: int,
    company_name: str = "NEOGAS",
    company_ruc: str = "",
    company_address: str = "",
    company_phone: str = "",
    is_reprint: bool = True,
) -> bytes:
    """Generate a PDF receipt mirroring the thermal shift-close ticket layout."""
    buf = io.BytesIO()

    # Narrow receipt width (80mm thermal style)
    page_w = 80 * mm
    page_h = 297 * mm  # long enough, will be trimmed by content

    doc = SimpleDocTemplate(
        buf, pagesize=(page_w, page_h),
        leftMargin=4 * mm, rightMargin=4 * mm,
        topMargin=5 * mm, bottomMargin=5 * mm,
    )

    styles = getSampleStyleSheet()

    # Custom styles for receipt look
    center_style = ParagraphStyle(
        'ReceiptCenter', parent=styles['Normal'],
        fontSize=7, alignment=1, leading=9,  # center
    )
    left_style = ParagraphStyle(
        'ReceiptLeft', parent=styles['Normal'],
        fontSize=7, alignment=0, leading=9,
    )
    bold_center = ParagraphStyle(
        'ReceiptBoldCenter', parent=center_style,
        fontSize=8, fontName='Helvetica-Bold', leading=10,
    )
    tiny_center = ParagraphStyle(
        'ReceiptTiny', parent=center_style,
        fontSize=6, leading=7, textColor=colors.grey,
    )
    amount_right = ParagraphStyle(
        'AmountRight', parent=styles['Normal'],
        fontSize=7, alignment=2, leading=9,
    )
    section_header = ParagraphStyle(
        'SectionHeader', parent=styles['Normal'],
        fontSize=7, fontName='Helvetica-Bold', leading=10,
        borderPadding=1,
    )

    elements = []

    def line(text: str, style=center_style):
        elements.append(Paragraph(text, style))

    def divider(char: str = "-"):
        elements.append(Paragraph(char * 42, tiny_center))

    def spacer(h: int = 2):
        elements.append(Spacer(1, h * mm))

    def row_pair(label: str, value: str):
        """Two-column row: label left, value right."""
        tbl = Table(
            [[Paragraph(label, left_style), Paragraph(value, amount_right)]],
            colWidths=[55 * mm, 17 * mm],
        )
        tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ]))
        elements.append(tbl)

    def row_pair_indent(label: str, value: str):
        """Indented row (for sub-items)."""
        tbl = Table(
            [[Paragraph(f"  {label}", left_style), Paragraph(value, amount_right)]],
            colWidths=[55 * mm, 17 * mm],
        )
        tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ]))
        elements.append(tbl)

    # ── Header ──
    line(f"<b>{company_name}</b>", bold_center)
    if company_ruc:
        line(f"RUC: {company_ruc}")
    if company_address:
        line(company_address)
    if company_phone:
        line(company_phone)

    divider()
    line("<b>CIERRE DE TURNO</b>", bold_center)
    if is_reprint:
        line("*** REIMPRESIÓN ***", bold_center)
    divider()

    line(f"Turno: #{shift_id}")
    line(f"Cajero: {cashier}")
    if opened_at:
        line(f"Apertura: {opened_at}")
    line(f"Cierre: {closed_at}")
    divider()

    # ── Cash Summary ──
    line("<b>RESUMEN DE EFECTIVO</b>", section_header)
    row_pair("Apertura", f"$ {opening_cash:,.2f}")

    if cash_income_count > 0:
        row_pair(f"Ingresos ({cash_income_count})", f"$ {cash_income:,.2f}")
    if cash_expense_count > 0:
        row_pair(f"Egresos ({cash_expense_count})", f"$ {cash_expense:,.2f}")
    if cash_deposits_count > 0:
        row_pair(f"Depósitos ({cash_deposits_count})", f"$ {cash_deposits:,.2f}")
    if cash_transfers_out_count > 0:
        row_pair(f"Transf. Env. ({cash_transfers_out_count})", f"$ {cash_transfers_out:,.2f}")
    if cash_transfers_in_count > 0:
        row_pair(f"Transf. Rec. ({cash_transfers_in_count})", f"$ {cash_transfers_in:,.2f}")
    if cash_safe_drops_count > 0:
        row_pair(f"Safe Drops ({cash_safe_drops_count})", f"$ {cash_safe_drops:,.2f}")

    row_pair(f"Ventas Efectivo ({sales_cash_count})", f"$ {sales_cash:,.2f}")

    divider()

    # ── Balance ──
    if surplus > 0:
        line(f"<b>Sobrante: $ {surplus:,.2f}</b>", bold_center)
    elif shortage > 0:
        line(f"<b>Faltante: $ {shortage:,.2f}</b>", bold_center)
    else:
        line("<b>Cuadre perfecto — $0.00</b>", bold_center)

    divider()

    # ── Non-Cash Sales ──
    if non_cash_sales:
        line("<b>OTRAS FORMAS DE PAGO</b>", section_header)
        for ncs in non_cash_sales:
            label = f"{ncs.get('method_name', '')} ({ncs.get('count', 0)})"
            total = float(ncs.get('total', 0))
            row_pair(label, f"$ {total:,.2f}")
        divider()

    # ── Totals ──
    line(f"Total Ventas: $ {total_sales:,.2f}")
    line(f"Despachos: {dispatch_count}")

    divider()
    line(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", tiny_center)

    doc.build(elements)
    return buf.getvalue()


def generate_shift_transactions_excel(
    shift_id: int,
    dispatches: list[dict],
    cash_movements: list[dict],
    company_name: str = "NEOGAS",
) -> bytes:
    """Generate Excel with 2 sheets: Despachos + Movimientos de Caja."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    title_font = Font(bold=True, size=12)
    subtitle_font = Font(size=9, color="666666")

    def write_sheet(ws, title_suffix: str, columns: list[str], rows: list[list]):
        ws.title = f"{title_suffix}"[:31]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        ws.cell(row=1, column=1, value=f"{company_name} — Turno #{shift_id} — {title_suffix}").font = title_font
        ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = subtitle_font

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.font = Font(size=9)

        # Auto-fit
        for col_idx in range(1, len(columns) + 1):
            max_len = len(columns[col_idx - 1])
            for row in rows:
                if col_idx <= len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = min(max_len + 4, 45)

        ws.freeze_panes = "A4"

    # ── Sheet 1: Dispatches ──
    ws1 = wb.active
    disp_cols = ["Orden", "Surtidor", "Lado", "Grado", "Cliente", "Placa",
                 "Pago", "Monto", "Volumen", "Precio Unit.", "IVA", "Estado", "SRI", "Hora"]
    disp_rows = []
    for d in dispatches:
        disp_rows.append([
            d.get("order_id", ""),
            str(d.get("dispenser_id", "")),
            d.get("side", ""),
            str(d.get("grade", "")),
            d.get("customer_name", "") or "",
            d.get("plate", "") or "",
            d.get("payment_method_name", "") or "",
            f"$ {float(d.get('final_amount', 0) or 0):,.2f}",
            str(d.get("final_volume", "") or ""),
            f"$ {float(d.get('unit_price', 0) or 0):,.4f}",
            f"$ {float(d.get('subsidy_amount', 0) or 0):,.2f}",
            d.get("status", ""),
            d.get("sri_status", "") or "PENDIENTE",
            d.get("created_at", "")[:19] if d.get("created_at") else "",
        ])
    write_sheet(ws1, "Despachos", disp_cols, disp_rows)

    # ── Sheet 2: Cash Movements ──
    ws2 = wb.create_sheet()
    cash_cols = ["ID", "Tipo", "Monto", "Observación", "Balance", "Usuario Rel.", "Hora"]
    cash_rows = []
    for cm in cash_movements:
        cash_rows.append([
            str(cm.get("movement_id", "")),
            cm.get("type", ""),
            f"$ {float(cm.get('amount', 0) or 0):,.2f}",
            cm.get("observation", "") or "",
            f"$ {float(cm.get('running_balance', 0) or 0):,.2f}",
            cm.get("related_user_name", "") or "",
            cm.get("created_at", "")[:19] if cm.get("created_at") else "",
        ])
    write_sheet(ws2, "Movimientos Caja", cash_cols, cash_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
